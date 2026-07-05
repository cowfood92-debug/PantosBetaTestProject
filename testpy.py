import streamlit as st
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.utils import column_index_from_string
import io
import re
from datetime import datetime

st.set_page_config(page_title="판토스 물류비 정산 시스템", layout="wide")

st.title("📦 판토스 마감 내역 정리")
st.write("마감내역서(PDF) 안의 모든 세부 비용 항목을 실제로 읽어서 엑셀 양식에 맞춰 정리해 드립니다.")

st.sidebar.header("📂 파일 업로드")
# 1번 엑셀 파일과 2번 PDF 파일 2개만으로 모든 기능을 구동합니다.
uploaded_excel = st.sidebar.file_uploader("1. 판토스 엑셀 양식 (기본 양식 및 공장입고 겸용)", type=["xlsx"])
uploaded_pantos = st.sidebar.file_uploader("2. 판토스 마감내역서 (PDF)", type=["pdf"])

# 기능 2에서 사용할 uploaded_factory 변수를 1번 엑셀 파일로 동일하게 연결합니다.
uploaded_factory = uploaded_excel

STANDARD_TRUCKING = 699000  # 철송 기준 내륙운송비 (580,000 + 119,000)
BASE_COMPONENT = 580000  # 안전운임제 기본 구성요소 (고정)
RAIL_COMPONENT = 119000  # 철송 이용 시 구성요소

# 마감내역서 FREIGHT 항목명 -> 최종 리포트 컬럼명 매핑
ITEM_MAP = {
    "OCEAN FREIGHT": "1. 해상운임 (OCEAN)",
    "WHARFAGE": "2. 화물입출항료 (WHARFAGE)",
    "CONTAINER CLEANING FEE": "3. 세척비 (CLEANING)",
    "CHASSIS CHARGE": "4. 샤시운임 (CHASSIS)",
    "EMERGENCY BUNKER SURCHARGE": "5. 유류할증료 (BUNKER)",
    "PREPULL CHARGE": "6. 프리풀 (PREPULL)",
    "TERMINAL HANDLING CHARGE": "7. 터미널조작비 (THC)",
    "TRUCKING CHARGE": "8. 트러킹 (TRUCKING)",
    "TRANSPORTATION CHARGE": "9. 선적지내륙운송 (TRANSPORT)",
    "DOCUMENT FEE": "10. 서류발급비 (DOC FEE)",
    "DOCUMENT FEE AT ORIGIN PORT": "11. 선적지서류비 (ORG DOC)",
}
OTHER_COL = "12. 기타 (그 외 항목)"
ITEM_COLS = list(ITEM_MAP.values()) + [OTHER_COL]

# 선적지(해외/출발지, USD 청구 항목) vs 도착지(국내/부산항, KRW 청구 항목) 비용 그룹
ORIGIN_COLS = [
    ITEM_MAP["OCEAN FREIGHT"],
    ITEM_MAP["CHASSIS CHARGE"],
    ITEM_MAP["EMERGENCY BUNKER SURCHARGE"],
    ITEM_MAP["PREPULL CHARGE"],
    ITEM_MAP["TRANSPORTATION CHARGE"],
    ITEM_MAP["DOCUMENT FEE AT ORIGIN PORT"],
]
DEST_COLS = [
    ITEM_MAP["WHARFAGE"],
    ITEM_MAP["CONTAINER CLEANING FEE"],
    ITEM_MAP["TERMINAL HANDLING CHARGE"],
    ITEM_MAP["TRUCKING CHARGE"],
    ITEM_MAP["DOCUMENT FEE"],
]

# 공장입고 엑셀의 "국내운송비(트러킹제외)"에 들어가는 도착지 항목 (트러킹 제외)
DEST_EXTRUCK_ITEMS = ["WHARFAGE", "CONTAINER CLEANING FEE", "TERMINAL HANDLING CHARGE", "DOCUMENT FEE"]

# 공장입고 엑셀의 선적지(외화 RATE) 항목 -> 컬럼 매핑에 쓰이는 원본 FREIGHT 명
ORIGIN_RATE_ITEMS = {
    "OCEAN FREIGHT": "N",
    "CHASSIS CHARGE": "O",
    "DOCUMENT FEE AT ORIGIN PORT": "P",
    "EMERGENCY BUNKER SURCHARGE": "Q",
    "PREPULL CHARGE": "R",
    "TRANSPORTATION CHARGE": "S",
}
KNOWN_FREIGHT_NAMES = set(ORIGIN_RATE_ITEMS.keys()) | set(DEST_EXTRUCK_ITEMS) | {"TRUCKING CHARGE"}


def to_num(s):
    if s is None:
        return None
    s = str(s).replace(",", "").strip()
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalize_lot(text):
    if text is None:
        return ""
    t = str(text)
    t = re.sub(r"\(.*?\)", "", t)
    t = t.replace("\n", " ")
    t = re.sub(r"\s+", "", t)
    t = t.rstrip(",")
    return t.upper()


@st.cache_data(show_spinner=False)
def parse_pantos_pdf(file_bytes):
    rows_all = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if any(r and "FREIGHT" in [str(c) for c in r] for r in table):
                    rows_all.extend(table)

    lots = {}
    order = []
    current_ref = None

    for r in rows_all:
        if r is None or len(r) < 18:
            continue
        if r[0] == "NO." or (r[0] and "T O T A L" in str(r[0])):
            continue

        if r[0] is not None and str(r[0]).strip().isdigit():
            current_ref = r[4]
            arr_dt_raw = r[1]
            arr_dt = None
            if arr_dt_raw:
                m = re.search(r"(\d{4}/\d{2}/\d{2})", str(arr_dt_raw))
                if m:
                    arr_dt = m.group(1)
            if current_ref not in lots:
                lots[current_ref] = {
                    "__EXRATE__": None,
                    "__TOTAL__": 0.0,
                    "__TRUCK_RATE__": None,
                    "__QTY__": None,
                    "__ARRDT__": arr_dt,
                    "__RATES__": {},
                }
                order.append(current_ref)

        freight_name = r[8]
        if current_ref is None or not freight_name or freight_name in ("FREIGHT",):
            continue

        per_type = r[5]
        qty = to_num(r[6])
        rate = to_num(r[9])
        curr = r[10]
        exrate = to_num(r[11])
        amt_l = to_num(r[13]) or 0
        proxy_l = to_num(r[16]) or 0
        sum_krw = to_num(r[17])

        value = amt_l if amt_l else proxy_l

        d = lots[current_ref]
        d[freight_name] = d.get(freight_name, 0) + value

        if freight_name not in d["__RATES__"]:
            d["__RATES__"][freight_name] = {"rate": rate, "curr": curr, "per_type": per_type, "qty": qty}

        if curr == "USD" and exrate and d["__EXRATE__"] is None:
            d["__EXRATE__"] = exrate
        if freight_name == "TRUCKING CHARGE" and rate:
            d["__TRUCK_RATE__"] = rate
        if per_type and per_type != "HBL" and qty:
            d["__QTY__"] = qty
        if sum_krw is not None:
            d["__TOTAL__"] = sum_krw

    return lots, order


def find_data_rows(ws):
    start_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, (int, float)) and v == 1:
            start_row = r
            break
    if start_row is None:
        return []

    rows = []
    r = start_row
    while True:
        v = ws.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            rows.append(r)
            r += 1
        else:
            break
    return rows


def fill_factory_excel(file_bytes, pdf_lots, pdf_order, sheet_name):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"'{sheet_name}' 시트를 찾을 수 없습니다.")

    ws = wb[sheet_name]
    warnings = []

    data_rows = find_data_rows(ws)
    if not data_rows:
        raise ValueError(f"'{sheet_name}' 시트에서 데이터 입력 행을 찾지 못했습니다.")

    n_template = len(data_rows)
    n_lots = len(pdf_order)

    if n_lots > n_template:
        warnings.append(
            f"⚠️ 마감내역서의 Lot 수({n_lots}건)가 엑셀 템플릿의 입력 가능 행수({n_template}행)보다 많습니다. 초과분은 입력되지 않았습니다."
        )

    w2 = ws["W2"].value
    x2 = ws["X2"].value
    w2 = w2 if isinstance(w2, (int, float)) else 200400
    x2 = x2 if isinstance(x2, (int, float)) else 119000

    def col(letter):
        return column_index_from_string(letter)

    for i, row in enumerate(data_rows):
        if i >= n_lots:
            break

        ref = pdf_order[i]
        d = pdf_lots[ref]
        rates = d.get("__RATES__", {})

        ws.cell(row=row, column=col("B")).value = ref

        arr_dt = d.get("__ARRDT__")
        if arr_dt:
            try:
                dt = datetime.strptime(arr_dt, "%Y/%m/%d")
                ws.cell(row=row, column=col("D")).value = dt
            except ValueError:
                ws.cell(row=row, column=col("D")).value = arr_dt

        if d.get("__EXRATE__"):
            ws.cell(row=row, column=col("M")).value = d["__EXRATE__"]

        for freight_name, col_letter in ORIGIN_RATE_ITEMS.items():
            info = rates.get(freight_name)
            if info and info.get("rate") is not None:
                r_rate = info["rate"]
                p_type = info.get("per_type", "")
                r_qty = info.get("qty")

                if p_type == "HBL":
                    calc_usd = r_rate * 1
                else:
                    calc_usd = r_rate * (r_qty if r_qty is not None else 1)

                ws.cell(row=row, column=col(col_letter)).value = calc_usd

        qty = d.get("__QTY__")
        if qty:
            ws.cell(row=row, column=col("AG")).value = qty
            ws.cell(row=row, column=col("AO")).value = qty

        truck_rate = d.get("__TRUCK_RATE__")
        if truck_rate:
            ws.cell(row=row, column=col("Y")).value = BASE_COMPONENT
            premium = truck_rate - BASE_COMPONENT
            if abs(truck_rate - STANDARD_TRUCKING) < 1:
                ws.cell(row=row, column=col("X")).value = x2 if abs(premium - x2) < 1 else premium
            else:
                ws.cell(row=row, column=col("W")).value = premium

        dest_sum = sum((d.get(name, 0) or 0) for name in DEST_EXTRUCK_ITEMS)
        if dest_sum:
            ws.cell(row=row, column=col("AA")).value = dest_sum

        l_cell = ws.cell(row=row, column=col("L"))
        if l_cell.value in (None, ""):
            l_cell.value = f"=SUM(N{row}:Q{row},R{row}:U{row})"

        v_cell = ws.cell(row=row, column=col("V"))
        if v_cell.value in (None, ""):
            v_cell.value = f"=L{row}*M{row}"

        z_cell = ws.cell(row=row, column=col("Z"))
        if z_cell.value in (None, ""):
            z_cell.value = f"=(Y{row}+W{row}+X{row})*AG{row}"

        ab_cell = ws.cell(row=row, column=col("AB"))
        if ab_cell.value in (None, ""):
            ab_cell.value = f"=SUM(Z{row}:AA{row})"

        ac_cell = ws.cell(row=row, column=col("AC"))
        if ac_cell.value in (None, ""):
            ac_cell.value = f"=V{row}+AB{row}"

    if "총계" in wb.sheetnames:
        ws_total = wb["총계"]
        header_row = 1
        for r in [1, 2, 3]:
            cells = [ws_total.cell(row=r, column=c).value for c in range(1, 15)]
            if any(c and any(k in str(c) for k in ["월", "구분", "항목", "내역"]) for c in cells):
                header_row = r
                break

        headers = {}
        for c in range(1, ws_total.max_column + 1):
            v = ws_total.cell(row=header_row, column=c).value
            if v:
                headers[str(v).strip()] = c

        target_total_row = None
        for r in range(header_row + 1, ws_total.max_row + 1):
            v = ws_total.cell(row=r, column=1).value
            if v and (sheet_name.replace("월", "") in str(v) or str(v).strip() == sheet_name.replace("월", "")):
                target_total_row = r
                break

        if target_total_row:
            total_qty_sum = sum(d.get("__QTY__", 0) or 0 for d in pdf_lots.values())
            total_krw_sum = sum(d.get("__TOTAL__", 0) or 0 for d in pdf_lots.values())

            total_usd_sum = 0
            for ref in pdf_order:
                d = pdf_lots[ref]
                usd_sum_for_lot = 0
                for name in ORIGIN_RATE_ITEMS:
                    info = d.get("__RATES__", {}).get(name, {})
                    r_rate = info.get("rate", 0) or 0
                    p_type = info.get("per_type", "")
                    r_qty = info.get("qty")

                    if p_type == "HBL":
                        usd_sum_for_lot += r_rate * 1
                    else:
                        usd_sum_for_lot += r_rate * (r_qty if r_qty is not None else 1)

                total_usd_sum += usd_sum_for_lot

            qty_filled, usd_filled, krw_filled = False, False, False
            for h_name, col_idx in headers.items():
                if any(k in h_name for k in ["물량", "수량", "컨테이너", "QTY", "Qty", "건수"]):
                    ws_total.cell(row=target_total_row, column=col_idx).value = total_qty_sum
                    qty_filled = True
                elif any(k in h_name for k in ["외화", "USD", "usd", "달러"]):
                    ws_total.cell(row=target_total_row, column=col_idx).value = total_usd_sum
                    usd_filled = True
                elif any(k in h_name for k in ["합계", "총액", "총금액", "금액", "원화", "TOTAL", "Total", "결제"]):
                    ws_total.cell(row=target_total_row, column=col_idx).value = total_krw_sum
                    krw_filled = True

            if not qty_filled and ws_total.max_column >= 2:
                ws_total.cell(row=target_total_row, column=2).value = total_qty_sum
            if not usd_filled and ws_total.max_column >= 3:
                ws_total.cell(row=target_total_row, column=3).value = total_usd_sum
            if not krw_filled and ws_total.max_column >= 4:
                ws_total.cell(row=target_total_row, column=4).value = total_krw_sum
        else:
            warnings.append(f"ℹ️ '총계' 시트는 존재하나 '{sheet_name}'에 해당하는 요약 행 데이터를 탐색하지 못했습니다.")

    return wb, warnings


pdf_lots, pdf_order = None, None
if uploaded_pantos:
    with st.spinner("마감내역서 PDF에서 모든 항목을 읽어오는 중..."):
        pdf_lots, pdf_order = parse_pantos_pdf(uploaded_pantos.getvalue())
    if not pdf_lots:
        st.error("마감내역서 PDF에서 데이터를 읽지 못했습니다. 파일 형식을 확인해주세요.")
        st.stop()
    st.success(f"마감내역서에서 총 {len(pdf_lots)}개 Lot의 세부 항목을 읽어왔습니다.")

# =========================================================================
# 기능 1. 마감내역서 데이터 기반 -> 항목별 상세정산 리포트 직접 표시
# =========================================================================
if uploaded_pantos:

    st.header("① 항목별 상세정산 리포트")

    st.subheader("🔍 세부 항목별 정산 내역 (미리보기)")

    results = []

    for lot_no_raw in pdf_order:
        d = pdf_lots[lot_no_raw]

        item_values = {c: 0 for c in ITEM_COLS}
        exrate = d.get("__EXRATE__")
        pdf_total = d.get("__TOTAL__", 0)
        truck_rate = d.get("__TRUCK_RATE__")
        qty = d.get("__QTY__")

        for name, amount in d.items():
            if name.startswith("__"):
                continue
            col_name = ITEM_MAP.get(name, OTHER_COL)
            item_values[col_name] += amount

        row_sum = sum(item_values.values())

        status = "정상"
        note = "특이사항 없음"
        is_road_trucking = False

        if truck_rate and truck_rate > STANDARD_TRUCKING:
            excess = truck_rate - STANDARD_TRUCKING
            is_road_trucking = True
            status = "🚚 육송운송 (내륙운송비 상승)"
            note = (
                f"철송 대신 육송운송이 적용되어 내륙운송비 단가가 상승했습니다. "
                f"(기본 {BASE_COMPONENT:,.0f}원 + {RAIL_COMPONENT + excess:,.0f}원 = {truck_rate:,.0f}원, "
                f"철송 기준 {STANDARD_TRUCKING:,.0f}원 대비 {excess:,.0f}원 상승)"
            )

        if pdf_total and abs(row_sum - pdf_total) > 1:
            status = "⚠️ 금액 불일치" if status == "정상" else status
            note = (note + " / " if note != "특이사항 없음" else "") + \
                   f"항목합계({row_sum:,.0f}) vs 명세서합계({pdf_total:,.0f}) 불일치"

        result_row = {
            "Lot 번호": lot_no_raw,
            "환율": f"{exrate:,.1f}" if exrate else "-",
        }
        for col_name in ITEM_COLS:
            result_row[col_name] = item_values[col_name]

        result_row["항목 합계"] = row_sum
        result_row["마감내역서 합계(원본)"] = pdf_total
        result_row["상태"] = status
        result_row["비고"] = note
        result_row["_컨테이너수"] = qty or 0
        result_row["_육송운송여부"] = is_road_trucking
        results.append(result_row)

    df_result = pd.DataFrame(results)

    if df_result.empty:
        empty_cols = ["Lot 번호", "환율"] + ITEM_COLS + ["항목 합계", "마감내역서 합계(원본)", "상태", "비고", "_컨테이너수", "_육송운송여부"]
        df_result = pd.DataFrame(columns=empty_cols)


    def highlight_status(val):
        if val == "🚚 육송운송 (내륙운송비 상승)":
            return "background-color: #ffe5cc; color: #b45f06; font-weight: bold;"
        if val == "❌ 매칭 실패":
            return "background-color: #999999; color: #ffffff; font-weight: bold;"
        if val == "⚠️ 금액 불일치":
            return "background-color: #fff3cd; color: #856404; font-weight: bold;"
        return ""


    df_display = df_result.drop(columns=["_컨테이너수", "_육송운송여부"], errors="ignore")
    st.dataframe(df_display.style.map(highlight_status, subset=["상태"]))

    total_pdf_invoice = sum(d.get("__TOTAL__", 0) for d in pdf_lots.values())
    total_report = df_result["항목 합계"].sum() if not df_result.empty else 0
    c1, c2 = st.columns(2)
    c1.metric("마감내역서 전체 합계", f"{total_pdf_invoice:,.0f} 원")
    c2.metric("리포트 항목 합계", f"{total_report:,.0f} 원", delta=f"{total_report - total_pdf_invoice:,.0f} 원 차이")

    st.subheader("📊 이번달 운송 분석")

    total_lots = len(df_result)
    total_qty = int(df_result["_컨테이너수"].sum()) if not df_result.empty else 0
    road_trucking_rows = df_result[df_result["_육송운송여부"] == True] if "_육송운송여부" in df_result.columns else pd.DataFrame()
    road_count = len(road_trucking_rows)

    m1, m2, m3 = st.columns(3)
    m1.metric("이번달 총 운송 건수", f"{total_lots} 건")
    m2.metric("총 컨테이너 수", f"{total_qty} 개")
    m3.metric("육송운송 발생 건수", f"{road_count} 건")

    origin_total = df_result[ORIGIN_COLS].sum().sum() if not df_result.empty else 0
    dest_total = df_result[DEST_COLS].sum().sum() if not df_result.empty else 0

    avg_origin = origin_total / total_lots if total_lots > 0 else 0
    avg_dest = dest_total / total_lots if total_lots > 0 else 0

    o1, o2 = st.columns(2)
    o1.metric(
        "선적지 운송비용 합계 (해상운임·샤시·프리풀 등)",
        f"{origin_total:,.0f} 원",
        f"건당 평균 {avg_origin:,.0f} 원",
    )
    o2.metric(
        "도착지 운송비용 합계 (부두사용료·터미널·트러킹 등)",
        f"{dest_total:,.0f} 원",
        f"건당 평균 {avg_dest:,.0f} 원",
    )

    st.markdown("**🔎 특이사항**")

    # --- 1. 육송운송 특이사항 ---
    if road_count > 0:
        road_extra_total = 0
        for _, r in road_trucking_rows.iterrows():
            m = re.search(r"([\d,]+)원\s*상승\)", r["비고"])
            if m:
                road_extra_total += to_num(m.group(1)) or 0
        st.warning(
            f"🚚 이번달 {road_count}건의 Lot에서 철송 대신 육송운송이 적용되어 내륙운송비가 상승했습니다 "
            f"(총 상승액 약 {road_extra_total:,.0f}원). 해당 Lot: "
            + ", ".join(road_trucking_rows["Lot 번호"].astype(str).tolist())
        )
    else:
        st.success("이번달은 육송운송으로 인한 내륙운송비 상승 사례가 없었습니다.")

    # --- 2. 유류할증료(BUNKER) 특이사항 추가 ---
    bunker_col = "5. 유류할증료 (BUNKER)"
    if not df_result.empty and bunker_col in df_result.columns:
        bunker_rows = df_result[df_result[bunker_col] > 0]
        bunker_count = len(bunker_rows)

        if bunker_count > 0:
            bunker_total = bunker_rows[bunker_col].sum()
            st.warning(
                f"⛽ 이번달 {bunker_count}건의 Lot에서 유류할증료가 추가로 발생했습니다 "
                f"(총 청구액 {bunker_total:,.0f}원). 해당 Lot: "
                + ", ".join(bunker_rows["Lot 번호"].astype(str).tolist())
            )

    mismatch_rows = df_result[df_result["상태"] == "⚠️ 금액 불일치"] if "상태" in df_result.columns else pd.DataFrame()
    if not mismatch_rows.empty:
        st.warning(
            f"⚠️ {len(mismatch_rows)}건의 Lot에서 항목 합계와 마감내역서 원본 합계가 일치하지 않습니다. "
            "엑셀 다운로드 파일에서 세부 내용을 확인해주세요: " + ", ".join(mismatch_rows["Lot 번호"].astype(str).tolist())
        )

    st.subheader("📥 전체 항목 분리 완료! 최종 엑셀 다운로드")

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        df_display.to_excel(writer, sheet_name="항목별_상세정산내역", index=False)

        workbook = writer.book
        worksheet = writer.sheets["항목별_상세정산내역"]

        for i, col_name in enumerate(df_display.columns):
            worksheet.set_column(i, i, 20)

        if "상태" in df_display.columns:
            status_col_idx = list(df_display.columns).index("상태")
            status_col_letter = chr(ord("A") + status_col_idx) if status_col_idx < 26 else "A"

            warning_format = workbook.add_format({"bg_color": "#FFE5CC", "font_color": "#B45F06"})
            worksheet.conditional_format(
                f"{status_col_letter}2:{status_col_letter}1000",
                {"type": "text", "criteria": "containing", "value": "육송운송", "format": warning_format},
            )
            mismatch_format = workbook.add_format({"bg_color": "#FFEB9C", "font_color": "#9C6500"})
            worksheet.conditional_format(
                f"{status_col_letter}2:{status_col_letter}1000",
                {"type": "text", "criteria": "containing", "value": "불일치", "format": mismatch_format},
            )

    st.download_button(
        label="📊 항목별 세부 정산 리포트 다운로드 (클릭)",
        data=buffer.getvalue(),
        file_name="판토스_항목별_세부리포트.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_report",
    )

# =========================================================================
# 기능 2. 판토스 공장입고 엑셀에 마감내역서 내용을 채워넣기
# =========================================================================
if uploaded_factory and uploaded_pantos:

    st.header("② 판토스 공장입고 엑셀 자동 채우기")
    st.write("공장입고 엑셀의 양식은 그대로 두고, 마감내역서에 나온 순서대로 Lot 정보를 채워 넣습니다. "
             "추가로 HBL/컨테이너 수량 비례 연산 및 환율 수식이 반영되며, 총계 시트의 데이터까지 빌드됩니다.")

    try:
        wb_preview = openpyxl.load_workbook(io.BytesIO(uploaded_factory.getvalue()), read_only=True)
        sheet_names = wb_preview.sheetnames
    except Exception as e:
        st.error(f"공장입고 엑셀을 여는 중 오류가 발생했습니다: {e}")
        sheet_names = []

    default_sheet = "6월" if "6월" in sheet_names else (sheet_names[0] if sheet_names else None)

    if default_sheet is None:
        st.error("공장입고 엑셀에서 시트를 찾지 못했습니다.")
    else:
        target_sheet = st.selectbox(
            "채워 넣을 시트를 선택하세요",
            options=sheet_names,
            index=sheet_names.index(default_sheet),
        )

        try:
            wb_filled, fill_warnings = fill_factory_excel(
                uploaded_factory.getvalue(), pdf_lots, pdf_order, target_sheet
            )
        except Exception as e:
            st.error(f"공장입고 엑셀을 채우는 중 오류가 발생했습니다: {e}")
            wb_filled = None
            fill_warnings = []

        if wb_filled is not None:
            for w in fill_warnings:
                if w.startswith("⚠️"):
                    st.warning(w)
                else:
                    st.info(w)

            st.success(f"'{target_sheet}' 시트 정밀 매핑 및 '총계' 요약 시트 연동 통합이 완료되었습니다.")

            out_buffer = io.BytesIO()
            wb_filled.save(out_buffer)

            st.download_button(
                label="📥 채워진 공장입고 엑셀 다운로드",
                data=out_buffer.getvalue(),
                file_name="판토스_공장입고_마감반영완료.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_factory",
            )